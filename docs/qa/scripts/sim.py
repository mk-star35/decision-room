# -*- coding: utf-8 -*-
"""Re-implement the V0.5 formula semantics in Python and check the numbers."""
import datetime as dt
import calendar as cal
import sys
import openpyxl

WB = sys.argv[1] if len(sys.argv) > 1 else "v05.xlsx"
wb = openpyxl.load_workbook(WB)
S = wb["Setup"]

CAT_FIRST, CAT_LAST = 20, 58
REC_FIRST, REC_LAST = 63, 90
ACC_FIRST, ACC_LAST = 12, 30
N_OCC = 53

YEAR = S["B4"].value
CURRENCY = S["B5"].value
PAYFREQ = S["B6"].value
PARTNER = S["B7"].value
FIRSTPAY = S["B8"].value
ASOF = dt.datetime(2026, 4, 30)      # stand-in for TODAY() so the run is reproducible

JAN1 = dt.datetime(YEAR, 1, 1)
DEC31 = dt.datetime(YEAR, 12, 31)

def eomonth(d, k=0):
    m = d.month - 1 + k
    y = d.year + m // 12
    m = m % 12 + 1
    return dt.datetime(y, m, cal.monthrange(y, m)[1])

def edate(d, k):
    m = d.month - 1 + k
    y = d.year + m // 12
    m = m % 12 + 1
    return dt.datetime(y, m, min(d.day, cal.monthrange(y, m)[1]))

def ceil_div(x):
    import math
    return -int(math.floor(-x))          # Excel: -INT(-x)

# --------------------------------------------------------------- accounts
accounts = []
for r in range(ACC_FIRST, ACC_LAST + 1):
    if S.cell(r, 1).value:
        accounts.append(dict(name=S.cell(r, 1).value, typ=S.cell(r, 2).value,
                             open=S.cell(r, 3).value or 0, limit=S.cell(r, 4).value or 0,
                             owner=S.cell(r, 5).value, active=S.cell(r, 6).value))
active_accounts = [a["name"] for a in accounts if a["active"] == "Yes"]

# ------------------------------------------------------------- categories
cats = []
for r in range(CAT_FIRST, CAT_LAST + 1):
    if S.cell(r, 1).value:
        cats.append(dict(main=S.cell(r, 1).value, sub=S.cell(r, 2).value,
                         grp=S.cell(r, 3).value, owner=S.cell(r, 4).value,
                         active=S.cell(r, 5).value))
act = [c for c in cats if c["active"] == "Yes"]

def main_of(sub):
    hit = [c["main"] for c in act if c["sub"] == sub]
    return hit[-1] if hit else None
def grp_of(main, sub):
    hit = [c["grp"] for c in act if c["main"] == main and c["sub"] == sub]
    return hit[-1] if hit else ""
def owner_of(main, sub):
    hit = [c["owner"] for c in act if c["main"] == main and c["sub"] == sub]
    return hit[-1] if hit else ""

# --------------------------------------------------------------- paydates
paydates = []
for k in range(69):
    d = None
    if PAYFREQ in ("Weekly", "Biweekly", "Every 4 weeks"):
        div = {"Weekly": 7, "Biweekly": 14, "Every 4 weeks": 28}[PAYFREQ]
        base = FIRSTPAY + dt.timedelta(days=ceil_div((JAN1 - FIRSTPAY).days / div) * div + k * div)
        d = base if base <= DEC31 else None
    elif PAYFREQ == "Monthly" and k < 12:
        m = k + 1
        d = dt.datetime(YEAR, m, min(FIRSTPAY.day, cal.monthrange(YEAR, m)[1]))
    elif PAYFREQ == "Semi-monthly" and k < 24:
        m = k // 2 + 1
        d = (dt.datetime(YEAR, m, min(FIRSTPAY.day, 15)) if k % 2 == 0
             else dt.datetime(YEAR, m, cal.monthrange(YEAR, m)[1]))
    if d:
        paydates.append(d)

# --------------------------------------------------------------- payments
recs = []
for r in range(REC_FIRST, REC_LAST + 1):
    if S.cell(r, 1).value and S.cell(r, 5).value:
        recs.append(dict(typ=S.cell(r, 1).value, sub=S.cell(r, 2).value,
                         amt=S.cell(r, 3).value or 0, freq=S.cell(r, 4).value,
                         start=S.cell(r, 5).value, end=S.cell(r, 6).value,
                         frm=S.cell(r, 7).value, to=S.cell(r, 8).value,
                         payer=S.cell(r, 9).value))

def occurrence(rc, k):
    st, f = rc["start"], rc["freq"]
    if f in ("Weekly", "Biweekly"):
        div = 7 if f == "Weekly" else 14
        return st + dt.timedelta(days=max(0, ceil_div((JAN1 - st).days / div)) * div + k * div)
    if f == "Monthly":
        return edate(st, max(0, (YEAR - st.year) * 12 + 1 - st.month) + k)
    if f in ("Quarterly", "Semiannual", "Annual"):
        step = {"Quarterly": 3, "Semiannual": 6, "Annual": 12}[f]
        diff = (YEAR - st.year) * 12 + 1 - st.month
        return edate(st, max(0, ceil_div(diff / step)) * step + k * step)
    if f == "Semi-monthly":
        if k >= 24:
            return None
        m = k // 2 + 1
        return (dt.datetime(YEAR, m, min(st.day, 15)) if k % 2 == 0
                else dt.datetime(YEAR, m, cal.monthrange(YEAR, m)[1]))
    if f == "One-time":
        return st if k == 0 else None
    return None

payments = []
for rc in recs:
    for k in range(N_OCC):
        d = occurrence(rc, k)
        if not d or not (JAN1 <= d <= DEC31) or d < rc["start"]:
            continue
        if rc["end"] and d > rc["end"]:
            continue
        main = main_of(rc["sub"]) or ("Bills" if rc["typ"] == "Bill" else rc["typ"])
        payments.append(dict(date=d, typ=rc["typ"], cat=main, sub=rc["sub"],
                             budget=rc["amt"], actual=rc["amt"], paid="Pending",
                             frm=rc["frm"] or "", to=rc["to"] or "",
                             owner=rc["payer"] or "", grp=grp_of(main, rc["sub"])))

# ------------------------------------------------------------- manual log
ML = wb["Manual Log"]
manual = []
for r in range(4, 501):
    if ML.cell(r, 1).value and ML.cell(r, 2).value:
        m, s_ = ML.cell(r, 2).value, ML.cell(r, 3).value
        manual.append(dict(date=ML.cell(r, 1).value, cat=m, sub=s_,
                           amt=ML.cell(r, 4).value or 0, acct=ML.cell(r, 5).value,
                           owner=owner_of(m, s_), grp=grp_of(m, s_)))

# ----------------------------------------------------------------- helpers
def pay_amt(p):
    return p["actual"] if p["paid"] == "Yes" else p["budget"]

def in_month(d, anchor):
    return anchor <= d < edate(anchor, 1)

def msum(pred):
    return sum(m["amt"] for m in manual if pred(m))

def psum(pred):
    return sum(p["actual"] for p in payments if pred(p))

BUCKETS = ("Bills", "Subscription", "Savings", "Debt")
def other_expense(pred):
    tot = msum(lambda m: pred(m) and m["cat"] != "Income")
    named = sum(msum(lambda m, b=b: pred(m) and m["cat"] == b) for b in BUCKETS)
    return tot - named

# ================================================================ REPORT
def line(k, v):
    print("  %-44s %s" % (k, v))

print("=" * 78)
print("V0.5 SIMULATION  (Budget Year %d, Pay %s, First Pay %s, As-of %s)"
      % (YEAR, PAYFREQ, FIRSTPAY.date(), ASOF.date()))
print("=" * 78)

print("\n[1] Pay dates")
line("count", len(paydates))
line("first / last", "%s / %s" % (paydates[0].date(), paydates[-1].date()))
line("PRE-PAY payments (should be 0)",
     sum(1 for p in payments if not any(d <= p["date"] for d in paydates)))
line("PRE-PAY manual rows (should be 0)",
     sum(1 for m in manual if not any(d <= m["date"] for d in paydates)))

print("\n[2] Payments ledger")
line("generated rows", len(payments))
line("blank From Account cells rendered as 0",
     sum(1 for p in payments if p["frm"] == 0 or p["to"] == 0))
line("distinct types", sorted(set(p["typ"] for p in payments)))

print("\n[3] Manual Log")
line("rows", len(manual))
bad = [m for m in manual
       if not any(c["main"] == m["cat"] and c["sub"] == m["sub"] for c in act)]
line("invalid category/sub pairs (QA_TEST expects 0)", len(bad))
for b in bad:
    line("   -> %s / %s" % (b["cat"], b["sub"]), b["date"].date())

print("\n[4] Accounts Tracker  (as of %s)" % ASOF.date())
print("  %-16s %10s %10s %10s %12s" % ("Account", "Start", "In", "Out", "End"))
for a in accounts:
    if a["active"] != "Yes":
        continue
    n = a["name"]
    inc = (psum(lambda p: p["to"] == n and p["date"] <= ASOF)
           + msum(lambda m: m["acct"] == n and m["cat"] == "Income" and m["date"] <= ASOF))
    outg = (psum(lambda p: p["frm"] == n and p["date"] <= ASOF)
            + msum(lambda m: m["acct"] == n and m["cat"] != "Income" and m["date"] <= ASOF))
    end = a["open"] + inc - outg
    print("  %-16s %10s %10s %10s %12s" % (n, a["open"], inc, outg, end))
    a["end"] = end

print("\n[5] Monthly Dashboard  (April 2026)")
A = dt.datetime(2026, 4, 1)
pm = lambda p: in_month(p["date"], A)
mm = lambda m: in_month(m["date"], A)
inc = psum(lambda p: pm(p) and p["typ"] == "Income") + msum(lambda m: mm(m) and m["cat"] == "Income")
rec = (psum(lambda p: pm(p) and p["typ"] not in ("Income", "Savings", "Debt", "Transfer"))
       + msum(lambda m: mm(m) and m["cat"] in ("Bills", "Subscription")))
var = other_expense(mm)
sav = psum(lambda p: pm(p) and p["typ"] == "Savings") + msum(lambda m: mm(m) and m["cat"] == "Savings")
deb = psum(lambda p: pm(p) and p["typ"] == "Debt") + msum(lambda m: mm(m) and m["cat"] == "Debt")
line("Income", inc); line("Recurring Expenses", rec); line("Variable & Other", var)
line("Savings", sav); line("Debt", deb); line("Net Cash Flow", inc - rec - var - sav - deb)
ml_out_total = msum(lambda m: mm(m) and m["cat"] != "Income")
named = rec - psum(lambda p: pm(p) and p["typ"] not in ("Income", "Savings", "Debt", "Transfer"))
line("cross-check: ML outflow = buckets + other",
     "%s == %s  %s" % (ml_out_total,
                       named + var + msum(lambda m: mm(m) and m["cat"] in ("Savings", "Debt")),
                       "OK" if abs(ml_out_total - (named + var + msum(lambda m: mm(m) and m["cat"] in ("Savings", "Debt")))) < 1e-9 else "MISMATCH"))

print("\n[6] Paycheck Dashboard 1  (period starting 2026-04-10 if present)")
start = next((d for d in paydates if d.month == 4), paydates[0])
nxt = next((d for d in paydates if d > start), DEC31)
pp = lambda p: start <= p["date"] < nxt
mp = lambda m: start <= m["date"] < nxt
i6 = psum(lambda p: pp(p) and p["typ"] == "Income") + msum(lambda m: mp(m) and m["cat"] == "Income")
b7 = (psum(lambda p: pp(p) and p["typ"] not in ("Income", "Savings", "Debt", "Transfer"))
      + msum(lambda m: mp(m) and m["cat"] in ("Bills", "Subscription")))
b8 = other_expense(mp)
b9 = psum(lambda p: pp(p) and p["typ"] == "Savings") + msum(lambda m: mp(m) and m["cat"] == "Savings")
b10 = psum(lambda p: pp(p) and p["typ"] == "Debt") + msum(lambda m: mp(m) and m["cat"] == "Debt")
line("period", "%s -> %s" % (start.date(), nxt.date()))
line("Income", i6); line("Recurring", b7); line("Variable & Other", b8)
line("Savings", b9); line("Debt", b10)
line("Projected End Balance (B11)", i6 - b7 - b8 - b9 - b10)
line("Safe to Spend (B12)", i6 - b7 - b9 - b10)
line("B11 == B12 ?", "IDENTICAL (bug)" if (i6 - b7 - b8 - b9 - b10) == (i6 - b7 - b9 - b10) and b8 != 0
     else "different  OK")

print("\n[7] 50/30/20  (April 2026)")
income = inc
for g in ("Needs", "Wants", "Savings"):
    amt = (msum(lambda m: mm(m) and m["grp"] == g)
           + psum(lambda p: pm(p) and p["grp"] == g))
    line(g, "%8s   %5.1f%% of income" % (amt, (amt / income * 100) if income else 0))
line("Take-home income", income)

print("\n[8] Savings goals  (as of %s)" % ASOF.date())
SV = wb["Savings-Sinking Funds"]
for r in range(4, 44):
    g = SV.cell(r, 1).value
    if not g:
        continue
    contrib = (psum(lambda p: p["sub"] == g and p["typ"] == "Savings" and p["date"] <= ASOF)
               + msum(lambda m: m["sub"] == g and m["cat"] == "Savings" and m["date"] <= ASOF))
    tgt, startbal = SV.cell(r, 2).value or 0, SV.cell(r, 3).value or 0
    curr = startbal + contrib
    line("%-16s target %6s" % (g, tgt),
         "contrib %6s  balance %7s  progress %5.1f%%"
         % (contrib, curr, (curr / tgt * 100) if tgt else 0))

print("\n[9] Debt payoff  (as of %s, method from sheet)" % ASOF.date())
DP = wb["Debt Payoff"]
method = DP["B3"].value
rows = []
for r in range(6, 16):
    n = DP.cell(r, 1).value
    if not n:
        continue
    b, apr = DP.cell(r, 2).value or 0, DP.cell(r, 3).value or 0
    pay = (DP.cell(r, 4).value or 0) + (DP.cell(r, 5).value or 0)
    paidsum = (psum(lambda p: p["sub"] == n and p["typ"] == "Debt" and p["date"] <= ASOF)
               + msum(lambda m: m["sub"] == n and m["cat"] == "Debt" and m["date"] <= ASOF))
    curr = max(0, b - paidsum)
    if pay <= 0 or curr <= 0:
        months = 0
    elif pay <= curr * apr / 12:
        months = "이자 초과"
    else:
        import math
        months = math.ceil(math.log(pay / (pay - curr * apr / 12)) / math.log(1 + apr / 12)) \
                 if apr else math.ceil(curr / pay)
    rows.append((n, b, apr, pay, curr, months))
line("method", method)
for n, b, apr, pay, curr, months in rows:
    line("%-14s start %6s apr %5.1f%%" % (n, b, apr * 100),
         "pay %5s  now %7s  est %s months" % (pay, curr, months))
order = sorted(rows, key=(lambda x: x[4]) if method == "Snowball" else (lambda x: -x[2]))
line("payoff order (%s)" % method, " > ".join(x[0] for x in order))

print("\n[10] Net Worth")
NW = wb["Net Worth Tracker"]
assets = 0
for r in range(5, 17):
    n, v = NW.cell(r, 1).value, NW.cell(r, 3).value
    if not n:
        continue
    val = next((a["end"] for a in accounts if a["name"] == n and "end" in a), None)
    if val is None:
        val = v if isinstance(v, (int, float)) else 0
    assets += val
    line("asset %-16s" % n, val)
liab = 0
for r in range(20, 31):
    n = NW.cell(r, 1).value
    if not n:
        continue
    acc = next((a["end"] for a in accounts if a["name"] == n and "end" in a), None)
    if acc is not None:
        val = abs(acc)
    else:
        val = next((x[4] for x in rows if x[0] == n), 0)
    liab += val
    line("liability %-12s" % n, val)
line("Total assets", assets); line("Total liabilities", liab)
line("NET WORTH", assets - liab)

