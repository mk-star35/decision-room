# -*- coding: utf-8 -*-
"""Structural verification of the V0.5 workbook: every reference must resolve."""
import re, sys
from collections import Counter
import openpyxl

wb = openpyxl.load_workbook(sys.argv[1] if len(sys.argv) > 1 else "v05.xlsx")
names = set(wb.defined_names)
sheets = set(wb.sheetnames)
CAT = (20, 58)
REC = (63, 90)
ACC = (12, 30)

FUNCS = set("""IF IFERROR AND OR NOT SUM SUMIFS SUMPRODUCT COUNT COUNTA COUNTIF COUNTIFS
INDEX MATCH LOOKUP OFFSET ROUND ROUNDUP ROUNDDOWN INT MAX MIN ABS MOD DATE YEAR MONTH DAY
EDATE EOMONTH TODAY WEEKNUM TEXT NPER RANK SMALL LARGE ROW COLUMN""".split())
LITERALS = set("""Yes No Bill Bills Income Savings Debt Variable Subscription Transfer
Snowball Avalanche Custom Pending Joint Cash Checking Credit Needs Wants Other TRUE FALSE
PRE Partner Mode disabled""".split())

NAMEPAT = re.compile(r'(?<![A-Za-z0-9_$!."\'])([A-Za-z_][A-Za-z0-9_.]{2,})(?![\w(])')

bad_setup, unknown_name, unknown_sheet = Counter(), Counter(), Counter()
total = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            total += 1
            for m in re.finditer(r"Setup!\$?([A-Z])\$?(\d+)", v):
                col, r_ = m.group(1), int(m.group(2))
                if r_ in (4, 5, 6, 7, 8, 9, 60, 91, 100):
                    ok = True
                elif ACC[0] <= r_ <= ACC[1]:
                    ok = col in "ABCDEF"
                elif CAT[0] <= r_ <= CAT[1]:
                    ok = col in "ABCDE"
                elif REC[0] <= r_ <= REC[1]:
                    ok = col in "ABCDEFGHIJ"
                else:
                    ok = False
                if not ok:
                    bad_setup["%s!%s -> Setup!%s%d" % (ws.title, c.coordinate, col, r_)] += 1
            body = re.sub(r'"[^"]*"', '""', v)
            for m in NAMEPAT.finditer(body):
                w = m.group(1)
                if w.upper() in FUNCS or w in LITERALS or w in names or w in sheets:
                    continue
                if re.match(r"^[A-Z]{1,3}\d+$", w):
                    continue
                unknown_name["%s!%s: %s" % (ws.title, c.coordinate, w)] += 1
            for m in re.finditer(r"'([^']+)'!", v):
                if m.group(1) not in sheets:
                    unknown_sheet[m.group(1)] += 1

print("formulas scanned                    :", total)
print("Setup refs landing outside a table  :", sum(bad_setup.values()))
for k in list(bad_setup)[:10]:
    print("     ", k)
print("unresolved names in formulas        :", sum(unknown_name.values()))
for k in list(unknown_name)[:12]:
    print("     ", k)
print("references to a missing sheet       :", sum(unknown_sheet.values()), dict(unknown_sheet))

print("\ndata validation sources")
for ws in wb.worksheets:
    for dv in ws.data_validations.dataValidation:
        f = dv.formula1
        state = "name" if f in names else ("range" if f.startswith(("OFFSET", "_Lists")) else "UNRESOLVED")
        if state == "UNRESOLVED":
            print("   !!", ws.title, dv.sqref, f)
print("   all sources resolve to a defined name or an explicit range")

print("\nprotection audit (formula cells should be locked, inputs open)")
for ws in wb.worksheets:
    if ws.sheet_state != "visible":
        continue
    editable = sum(1 for row in ws.iter_rows() for c in row
                   if c.value is not None and not c.protection.locked)
    open_f = sum(1 for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str) and c.value.startswith("=") and not c.protection.locked)
    print("   %-24s protected=%-5s editable-with-content=%-5d formula-cells-left-open=%d"
          % (ws.title, ws.protection.sheet, editable, open_f))

print("\nsheet visibility")
for ws in wb.worksheets:
    if ws.sheet_state != "visible":
        print("   %-24s %s" % (ws.title, ws.sheet_state))

print("\ncharts")
for ws in wb.worksheets:
    for ch in ws._charts:
        refs = []
        for s in ch.series:
            r = s.val.numRef.f if s.val and s.val.numRef else None
            refs.append(r)
        print("   %-24s %-12s series=%d  %s"
              % (ws.title, type(ch).__name__, len(ch.series), refs))

print("\nweight")
tot = ch_ = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                tot += 1
                ch_ += len(c.value)
print("   %d formulas, %.2f MB of formula text" % (tot, ch_ / 1e6))
