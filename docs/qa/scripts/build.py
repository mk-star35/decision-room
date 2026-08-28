#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Etsy Budget Benchmark V0.4 -> V0.5 QA fix build."""
import copy, re, shutil, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

SRC = sys.argv[1]
DST = sys.argv[2]
shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)

BLUE   = "FF0000FF"   # user input
GREEN  = "FF008000"   # link from another sheet
BLACK  = "FF000000"   # formula
PURPLE = "FF7030A0"   # control
ORANGE = "FFFFC000"

# =====================================================================
# 0.  layout constants
# =====================================================================
CAT_FIRST, CAT_LAST = 20, 58          # Setup category data rows (was 20..30)
REC_TITLE           = 61              # was 31
REC_HDR             = 62              # was 32
REC_FIRST, REC_LAST = 63, 90          # was 33..60
N_ITEMS             = REC_LAST - REC_FIRST + 1        # 28
N_OCC               = 53
ACC_FIRST, ACC_LAST = 12, 30
LIST_CAT_ROWS       = CAT_LAST - CAT_FIRST + 1        # 39
ASOF                = "Setup!$B$9"
PAYD                = "_Lists!$W$2:$W$70"   # plain range: OFFSET names are volatile

setup = wb["Setup"]

# =====================================================================
# 1.  Setup — move the RECURRING block down, add As of Date
# =====================================================================
# 1a. move rows 31..60 -> 61..90 (values + styles), bottom-up
setup.unmerge_cells("A31:J31")
for old in range(60, 30, -1):
    new = old + 30
    for col in range(1, 15):
        s = setup.cell(old, col)
        d = setup.cell(new, col)
        d.value = s.value
        d._style = copy.copy(s._style)
        s.value = None
    if old in setup.row_dimensions:
        setup.row_dimensions[new].height = setup.row_dimensions[old].height

setup.merge_cells("A%d:J%d" % (REC_TITLE, REC_TITLE))

# 1b. restore a neutral style on the vacated rows and extend the category grid
for r in range(29, CAT_LAST + 1):
    for col in range(1, 6):
        src = setup.cell(28, col)
        d = setup.cell(r, col)
        d._style = copy.copy(src._style)
        d.value = None

# 1c. As of Date control
setup["A9"] = "As of Date"
setup["A9"]._style = copy.copy(setup["A8"]._style)
setup["B9"] = "=TODAY()"
setup["B9"]._style = copy.copy(setup["B8"]._style)
setup["B9"].font = Font(name="Arial", size=11, bold=True, color=PURPLE)
setup["B9"].number_format = "yyyy-mm-dd"
setup["C9"] = '=IF(B9="","","잔액·진행률 시트는 이 날짜까지의 거래만 집계합니다")'
setup["C9"]._style = copy.copy(setup["C8"]._style)

setup["C8"] = ('=IF(B8="","","Budget Year 1월 1일 이전 날짜를 권장합니다. '
               '급여일은 이 날짜를 기준으로 앞뒤로 생성됩니다")')

# 1d. demo data: pay cycle starts at the top of the budget year
setup["B8"] = __import__("datetime").datetime(2026, 1, 2)

# 1d-2. demo balances that do not read as "already finished" on day one
setup["C14"] = -2400          # Credit Card opening balance

# 1e. demo categories the other sheets rely on (P1-09)
NEW_CATS = [
    ("Savings",      "Vacation",      "Savings", "Joint",     "Yes"),
    ("Savings",      "Laptop",        "Savings", "Partner 1", "Yes"),
    ("Debt",         "Car Loan",      "Needs",   "Joint",     "Yes"),
    ("Debt",         "Student Loan",  "Needs",   "Partner 2", "Yes"),
    ("Variable",     "Transport",     "Needs",   "Joint",     "Yes"),
]
for i, row in enumerate(NEW_CATS):
    r = 29 + i
    for c, v in enumerate(row, start=1):
        setup.cell(r, c).value = v
        setup.cell(r, c).font = Font(name="Arial", size=11, color=BLUE)

# 1f. recurring block: warning + Day/Rule header rename
setup.cell(REC_HDR, 10).value = "Note (참고용)"
# overflow warnings live OUTSIDE the range they count, or COUNTA would see themselves
setup["L61"] = ('=IF(COUNTA($A$%d:$A$100)>0,"⚠ 반복거래는 %d건까지 지원됩니다. '
                '%d행 아래 입력은 집계되지 않습니다","")' % (REC_LAST + 1, N_ITEMS, REC_LAST))
setup["L61"].font = Font(name="Arial", size=10, bold=True, color="FFC00000")
setup["G59"] = ('=IF(COUNTA($A$%d:$A$60)>0,"⚠ 카테고리는 %d행까지 입력하세요. '
                '아래 입력은 집계되지 않습니다","")' % (CAT_LAST + 1, CAT_LAST))
setup["G59"].font = Font(name="Arial", size=10, bold=True, color="FFC00000")

# =====================================================================
# 2.  _Lists — helper ranks, grouped category list, pay dates
# =====================================================================
L = wb["_Lists"]

L["B7"] = None                      # drop "Custom" pay frequency (P1-13)

hdr_style = copy.copy(L["A1"]._style)
def head(col, text):
    c = L.cell(1, col); c.value = text; c._style = copy.copy(hdr_style)

# Y: static base list of main categories (P2-06)
head(25, "MainCategoryBase")
for i, v in enumerate(["Income", "Bills", "Variable", "Subscription",
                       "Savings", "Debt", "Transfer", "Other"]):
    L.cell(2 + i, 25).value = v

CAT_A = "Setup!$A$%d:$A$%d" % (CAT_FIRST, CAT_LAST)
CAT_B = "Setup!$B$%d:$B$%d" % (CAT_FIRST, CAT_LAST)
CAT_E = "Setup!$E$%d:$E$%d" % (CAT_FIRST, CAT_LAST)
ACC_A = "Setup!$A$%d:$A$%d" % (ACC_FIRST, ACC_LAST)
ACC_F = "Setup!$F$%d:$F$%d" % (ACC_FIRST, ACC_LAST)
LAST_L = 1 + LIST_CAT_ROWS          # 40
LAST_ACC_L = 1 + (ACC_LAST - ACC_FIRST + 1)   # 20

# AE / AF / AG : first-occurrence sequence numbers (forward order, no array formulas)
head(31, "MainCatRank"); head(32, "AccountRank"); head(33, "SubcatRank")
for n in range(1, LIST_CAT_ROWS + 1):
    r  = 1 + n                       # _Lists row
    sr = CAT_FIRST + n - 1           # Setup row
    if n == 1:
        L.cell(r, 31).value = '=IF(OR(Setup!$A$%d="",Setup!$E$%d<>"Yes"),"",1)' % (sr, sr)
        L.cell(r, 33).value = '=IF(OR(Setup!$B$%d="",Setup!$E$%d<>"Yes"),"",1)' % (sr, sr)
    else:
        L.cell(r, 31).value = (
            '=IF(OR(Setup!$A$%d="",Setup!$E$%d<>"Yes"),"",'
            'IF(COUNTIFS(Setup!$A$%d:$A$%d,Setup!$A$%d,Setup!$E$%d:$E$%d,"Yes")>0,"",'
            'MAX($AE$1:AE%d)+1))' % (sr, sr, CAT_FIRST, sr - 1, sr, CAT_FIRST, sr - 1, r - 1))
        L.cell(r, 33).value = (
            '=IF(OR(Setup!$B$%d="",Setup!$E$%d<>"Yes"),"",'
            'IF(COUNTIFS(Setup!$B$%d:$B$%d,Setup!$B$%d,Setup!$E$%d:$E$%d,"Yes")>0,"",'
            'MAX($AG$1:AG%d)+1))' % (sr, sr, CAT_FIRST, sr - 1, sr, CAT_FIRST, sr - 1, r - 1))
for n in range(1, ACC_LAST - ACC_FIRST + 2):
    r  = 1 + n
    sr = ACC_FIRST + n - 1
    if n == 1:
        L.cell(r, 32).value = '=IF(OR(Setup!$A$%d="",Setup!$F$%d<>"Yes"),"",1)' % (sr, sr)
    else:
        L.cell(r, 32).value = (
            '=IF(OR(Setup!$A$%d="",Setup!$F$%d<>"Yes"),"",'
            'IF(COUNTIFS(Setup!$A$%d:$A$%d,Setup!$A$%d,Setup!$F$%d:$F$%d,"Yes")>0,"",'
            'MAX($AF$1:AF%d)+1))' % (sr, sr, ACC_FIRST, sr - 1, sr, ACC_FIRST, sr - 1, r - 1))

# N / O / P : the visible dynamic lists, now in entry order (P2-05)
for r in range(2, 43):
    L.cell(r, 14).value = None; L.cell(r, 16).value = None
    for col in range(17, 23):
        L.cell(r, col).value = None
for r in range(2, 21):
    L.cell(r, 15).value = None
for n in range(1, LIST_CAT_ROWS + 1):
    r = 1 + n
    L.cell(r, 14).value = '=IFERROR(INDEX(%s,MATCH(%d,$AE$2:$AE$%d,0)),"")' % (CAT_A, n, LAST_L)
    L.cell(r, 16).value = '=IFERROR(INDEX(%s,MATCH(%d,$AG$2:$AG$%d,0)),"")' % (CAT_B, n, LAST_L)
for n in range(1, ACC_LAST - ACC_FIRST + 2):
    r = 1 + n
    L.cell(r, 15).value = '=IFERROR(INDEX(%s,MATCH(%d,$AF$2:$AF$%d,0)),"")' % (ACC_A, n, LAST_ACC_L)
head(17, ""); head(18, ""); head(19, ""); head(20, ""); head(21, ""); head(22, "")

# AA / AB / AC : category list grouped by main category -> dependent dropdown (P0-02)
head(27, "CatSortKey"); head(28, "GroupedMain"); head(29, "GroupedSub")
for n in range(1, LIST_CAT_ROWS + 1):
    r  = 1 + n
    sr = CAT_FIRST + n - 1
    L.cell(r, 27).value = (
        '=IF(OR(Setup!$B$%d="",Setup!$E$%d<>"Yes"),"",'
        'IFERROR(MATCH(Setup!$A$%d,$N$2:$N$%d,0),50)*1000+%d)' % (sr, sr, sr, LAST_L, n))
    L.cell(r, 28).value = '=IFERROR(INDEX(%s,MOD(SMALL($AA$2:$AA$%d,%d),1000)),"")' % (CAT_A, LAST_L, n)
    L.cell(r, 29).value = '=IFERROR(INDEX(%s,MOD(SMALL($AA$2:$AA$%d,%d),1000)),"")' % (CAT_B, LAST_L, n)

# AH / AI : expense sub-categories, feeds the Budget vs Actual block (P1-04)
head(34, "ExpenseSubRank"); head(35, "ExpenseSub")
for n in range(1, LIST_CAT_ROWS + 1):
    r = 1 + n
    if n == 1:
        L.cell(r, 34).value = '=IF(OR($AB2="",$AB2="Income"),"",1)'
    else:
        L.cell(r, 34).value = ('=IF(OR($AB%d="",$AB%d="Income"),"",MAX($AH$1:AH%d)+1)'
                               % (r, r, r - 1))
    L.cell(r, 35).value = ('=IFERROR(INDEX($AC$2:$AC$%d,MATCH(%d,$AH$2:$AH$%d,0)),"")'
                           % (LAST_L, n, LAST_L))

# W : pay dates, now generated backwards from the reference date too (P1-12)
def paydate(k):
    def slide(div):
        base = ('Setup!$B$8-INT(-((DATE(Setup!$B$4,1,1)-Setup!$B$8)/%d))*%d+%d*%d'
                % (div, div, k, div))
        return 'IF(%s<=DATE(Setup!$B$4,12,31),%s,"")' % (base, base)
    mon = ('IF(%d<12,DATE(Setup!$B$4,%d,MIN(DAY(Setup!$B$8),'
           'DAY(EOMONTH(DATE(Setup!$B$4,%d,1),0)))),"")' % (k, k + 1, k + 1))
    semi = ('IF(%d<24,IF(MOD(%d,2)=0,DATE(Setup!$B$4,%d,MIN(DAY(Setup!$B$8),15)),'
            'DATE(Setup!$B$4,%d,DAY(EOMONTH(DATE(Setup!$B$4,%d,1),0)))),"")'
            % (k, k, k // 2 + 1, k // 2 + 1, k // 2 + 1))
    return ('=IF(Setup!$B$8="","",'
            'IF(Setup!$B$6="Weekly",%s,'
            'IF(Setup!$B$6="Biweekly",%s,'
            'IF(Setup!$B$6="Every 4 weeks",%s,'
            'IF(Setup!$B$6="Monthly",%s,'
            'IF(Setup!$B$6="Semi-monthly",%s,""))))))'
            % (slide(7), slide(14), slide(28), mon, semi))
for k in range(0, 69):
    L.cell(2 + k, 23).value = paydate(k)

# defined names
DN = {
    "MainCategory":           'OFFSET(_Lists!$N$2,0,0,MAX(1,SUMPRODUCT(--(_Lists!$N$2:$N$%d<>""))),1)' % LAST_L,
    "ActiveAccountNames":     'OFFSET(_Lists!$O$2,0,0,MAX(1,SUMPRODUCT(--(_Lists!$O$2:$O$%d<>""))),1)' % LAST_ACC_L,
    "ActiveSubcategoryNames": 'OFFSET(_Lists!$P$2,0,0,MAX(1,SUMPRODUCT(--(_Lists!$P$2:$P$%d<>""))),1)' % LAST_L,
    "PayFrequency":           '_Lists!$B$2:$B$6',
    "MainCategoryBase":       '_Lists!$Y$2:$Y$9',
    "PayDates":               'OFFSET(_Lists!$W$2,0,0,MAX(1,COUNT(_Lists!$W$2:$W$70)),1)',
}
for name, val in DN.items():
    if name in wb.defined_names:
        wb.defined_names[name].value = val
    else:
        wb.defined_names.add(DefinedName(name, attr_text=val))
for dead in ["Income", "Bills", "Variable", "Savings", "Debt", "Subscription",
             "CategoryNames", "SubcategoryNames", "AccountNames", "PayPeriod"]:
    if dead in wb.defined_names:
        del wb.defined_names[dead]

# =====================================================================
# 3.  Payments — regenerate every row, compactly (P0-07, P1-18, P3-01)
# =====================================================================
pay = wb["Payments"]
for r in range(4, 1501):
    for c in range(1, 15):
        pay.cell(r, c).value = None

pay["N3"] = "Occurrence date (helper)"
pay["N3"]._style = copy.copy(pay["M3"]._style)
pay["K3"] = "Pay Period (연 누적)"

def occurrence(S, k):
    """raw occurrence date for Setup row S, occurrence k"""
    E = "Setup!$E$%d" % S
    D = "Setup!$D$%d" % S
    Y1 = "DATE(Setup!$B$4,1,1)"
    def slide(div):
        return '%s+MAX(0,-INT(-((%s-%s)/%d)))*%d+%d*%d' % (E, Y1, E, div, div, k, div)
    diff = '(Setup!$B$4-YEAR(%s))*12+1-MONTH(%s)' % (E, E)
    def edate(step):
        if step == 1:
            return 'EDATE(%s,MAX(0,%s)+%d)' % (E, diff, k)
        return 'EDATE(%s,MAX(0,-INT(-((%s)/%d)))*%d+%d*%d)' % (E, diff, step, step, k, step)
    semi = ('IF(%d<24,IF(MOD(%d,2)=0,DATE(Setup!$B$4,%d,MIN(DAY(%s),15)),'
            'DATE(Setup!$B$4,%d,DAY(EOMONTH(DATE(Setup!$B$4,%d,1),0)))),"")'
            % (k, k, k // 2 + 1, E, k // 2 + 1, k // 2 + 1))
    return ('=IF(OR(Setup!$B$%d="",%s=""),"",IFERROR('
            'IF(%s="Weekly",%s,'
            'IF(%s="Biweekly",%s,'
            'IF(%s="Monthly",%s,'
            'IF(%s="Quarterly",%s,'
            'IF(%s="Semiannual",%s,'
            'IF(%s="Annual",%s,'
            'IF(%s="Semi-monthly",%s,'
            'IF(%s="One-time",IF(%d=0,%s,""),""))))))))'
            ',""))'
            % (S, E,
               D, slide(7), D, slide(14), D, edate(1), D, edate(3),
               D, edate(6), D, edate(12), D, semi, D, k, E))

thin = Side(style="thin", color="FFD9D9D9")
for i in range(N_ITEMS):
    S = REC_FIRST + i
    for k in range(N_OCC):
        r = 4 + i * N_OCC + k
        pay.cell(r, 14).value = occurrence(S, k)
        pay.cell(r, 1).value = (
            '=IF($N%d="","",IF(AND($N%d>=DATE(Setup!$B$4,1,1),$N%d<=DATE(Setup!$B$4,12,31),'
            '$N%d>=Setup!$E$%d,OR(Setup!$F$%d="",$N%d<=Setup!$F$%d)),$N%d,""))'
            % (r, r, r, r, S, S, r, S, r))
        pay.cell(r, 2).value = '=IF($A%d="","",Setup!$A$%d)' % (r, S)
        r0 = 4 + i * N_OCC
        if k == 0:
            pay.cell(r, 3).value = (
                '=IF($A%d="","",IFERROR(LOOKUP(2,1/((%s=Setup!$B$%d)*(%s="Yes")),%s),'
                'IF(Setup!$A$%d="Bill","Bills",Setup!$A$%d)))'
                % (r, CAT_B, S, CAT_E, CAT_A, S, S))
        else:
            pay.cell(r, 3).value = '=IF($A%d="","",$C$%d)' % (r, r0)
        pay.cell(r, 4).value = '=IF($A%d="","",Setup!$B$%d)' % (r, S)
        pay.cell(r, 5).value = '=IF($A%d="","",Setup!$C$%d)' % (r, S)
        pay.cell(r, 6).value = '=IF($E%d="","",$E%d)' % (r, r)
        pay.cell(r, 7).value = '=IF($A%d="","","Pending")' % r
        pay.cell(r, 8).value = '=IF($A%d="","",IF(Setup!$G$%d="","",Setup!$G$%d))' % (r, S, S)
        pay.cell(r, 9).value = '=IF($A%d="","",IF(Setup!$H$%d="","",Setup!$H$%d))' % (r, S, S)
        pay.cell(r, 10).value = '=IF($A%d="","",IF(Setup!$I$%d="","",Setup!$I$%d))' % (r, S, S)
        pay.cell(r, 11).value = ('=IF($A%d="","",IF(COUNTIF(' + PAYD + ',"<="&$A%d)=0,"PRE-PAY",'
                                 '"P"&COUNTIF(' + PAYD + ',"<="&$A%d)))') % (r, r, r)
        if k == 0:
            pay.cell(r, 13).value = (
                '=IF($A%d="","",IFERROR(LOOKUP(2,1/((%s=$C%d)*(%s=$D%d)*(%s="Yes")),'
                'Setup!$C$%d:$C$%d),""))' % (r, CAT_A, r, CAT_B, r, CAT_E, CAT_FIRST, CAT_LAST))
        else:
            pay.cell(r, 13).value = '=IF($A%d="","",$M$%d)' % (r, r0)
        nc = pay.cell(r, 14)
        nc.number_format = "yyyy-mm-dd"
        nc.font = Font(name="Arial", size=10, color="FFBFBFBF")
pay.column_dimensions["N"].hidden = True
pay.column_dimensions["M"].hidden = False          # 50/30/20 Group is real data (P0-03)

# =====================================================================
# 4.  Manual Log — kill the 40 helper columns, fix the dependent list
# =====================================================================
ml = wb["Manual Log"]
for r in range(1, 501):
    for c in range(13, 53):                 # M..AZ
        ml.cell(r, c).value = None
for c in range(13, 53):
    ml.column_dimensions[get_column_letter(c)].hidden = False
for r in range(4, 501):
    ml.cell(r, 8).value = ('=IF($A%d="","",IF(COUNTIF(%s,"<="&$A%d)=0,"PRE-PAY",'
                           '"P"&COUNTIF(%s,"<="&$A%d)))' % (r, PAYD, r, PAYD, r))
    ml.cell(r, 6).value = ('=IF(OR($B%d="",$C%d=""),"",IFERROR(LOOKUP(2,1/((%s=$B%d)*(%s=$C%d)*'
                           '(%s="Yes")),Setup!$D$%d:$D$%d),""))'
                           % (r, r, CAT_A, r, CAT_B, r, CAT_E, CAT_FIRST, CAT_LAST))
    ml.cell(r, 7).value = ('=IF(OR($B%d="",$C%d=""),"",IFERROR(LOOKUP(2,1/((%s=$B%d)*(%s=$C%d)*'
                           '(%s="Yes")),Setup!$C$%d:$C$%d),""))'
                           % (r, r, CAT_A, r, CAT_B, r, CAT_E, CAT_FIRST, CAT_LAST))

# demo data: drop the QA junk (P0-01)
for r in range(9, 26):
    for c in range(1, 6):
        ml.cell(r, c).value = None
DEMO = [
    ("2026-04-15", "Bills",        "Utilities",  38,  "Main Checking"),
    ("2026-04-18", "Variable",     "Groceries",  110, "Credit Card"),
    ("2026-04-20", "Variable",     "Dining Out", 60,  "Credit Card"),
    ("2026-04-21", "Variable",     "Transport",  46,  "Main Checking"),
    ("2026-04-25", "Variable",     "Shopping",   72,  "Credit Card"),
]
import datetime as _dt
for i, (d, mc, sc, amt, acct) in enumerate(DEMO):
    r = 9 + i
    ml.cell(r, 1).value = _dt.datetime.strptime(d, "%Y-%m-%d")
    ml.cell(r, 2).value = mc
    ml.cell(r, 3).value = sc
    ml.cell(r, 4).value = amt
    ml.cell(r, 5).value = acct

# =====================================================================
# 5.  shared SUMIFS fragments
# =====================================================================
P_AMT   = "Payments!$F$4:$F$1500"
P_BUD   = "Payments!$E$4:$E$1500"
P_DATE  = "Payments!$A$4:$A$1500"
P_TYPE  = "Payments!$B$4:$B$1500"
P_SUB   = "Payments!$D$4:$D$1500"
P_PAID  = "Payments!$G$4:$G$1500"
P_FROM  = "Payments!$H$4:$H$1500"
P_TO    = "Payments!$I$4:$I$1500"
P_OWN   = "Payments!$J$4:$J$1500"
P_GRP   = "Payments!$M$4:$M$1500"
M_AMT   = "'Manual Log'!$D$4:$D$500"
M_DATE  = "'Manual Log'!$A$4:$A$500"
M_CAT   = "'Manual Log'!$B$4:$B$500"
M_SUB   = "'Manual Log'!$C$4:$C$500"
M_OWN   = "'Manual Log'!$F$4:$F$500"
M_GRP   = "'Manual Log'!$G$4:$G$500"

def mlsum(period, *crit):
    """Manual Log SUMIFS over a period; period is a list of (range, test)"""
    parts = ",".join("%s,%s" % (rg, tt) for rg, tt in period)
    extra = ",".join("%s,%s" % (rg, tt) for rg, tt in crit)
    return "SUMIFS(%s,%s%s)" % (M_AMT, parts, ("," + extra) if extra else "")

def month_period(anchor):
    return [(M_DATE, '">="&%s' % anchor), (M_DATE, '"<"&EDATE(%s,1)' % anchor)]

def month_period_p(anchor):
    return [(P_DATE, '">="&%s' % anchor), (P_DATE, '"<"&EDATE(%s,1)' % anchor)]

def paysum(period, *crit):
    parts = ",".join("%s,%s" % (rg, tt) for rg, tt in period)
    extra = ",".join("%s,%s" % (rg, tt) for rg, tt in crit)
    return "SUMIFS(%s,%s%s)" % (P_AMT, parts, ("," + extra) if extra else "")

def other_expense(period):
    """every Manual Log outflow that is not one of the five named buckets (P1-02)"""
    return ("%s-%s-%s-%s-%s"
            % (mlsum(period, (M_CAT, '"<>Income"')),
               mlsum(period, (M_CAT, '"Bills"')),
               mlsum(period, (M_CAT, '"Subscription"')),
               mlsum(period, (M_CAT, '"Savings"')),
               mlsum(period, (M_CAT, '"Debt"'))))

# =====================================================================
# 6.  Accounts Tracker — as-of cutoff + Paid?-aware amounts (P1-01)
# =====================================================================
at = wb["Accounts Tracker"]
for r in range(4, 23):
    a = "$A%d" % r
    at.cell(r, 4).value = (
        '=IF(%s="","",SUMIFS(%s,%s,%s,%s,"<="&%s,%s,"Yes")'
        '+SUMIFS(%s,%s,%s,%s,"<="&%s,%s,"<>Yes")'
        '+SUMIFS(%s,%s,%s,%s,"Income",%s,"<="&%s))'
        % (a, P_AMT, P_TO, a, P_DATE, ASOF, P_PAID,
           P_BUD, P_TO, a, P_DATE, ASOF, P_PAID,
           M_AMT, "'Manual Log'!$E$4:$E$500", a, M_CAT, M_DATE, ASOF))
    at.cell(r, 5).value = (
        '=IF(%s="","",SUMIFS(%s,%s,%s,%s,"<="&%s,%s,"Yes")'
        '+SUMIFS(%s,%s,%s,%s,"<="&%s,%s,"<>Yes")'
        '+SUMIFS(%s,%s,%s,%s,"<>Income",%s,"<="&%s))'
        % (a, P_AMT, P_FROM, a, P_DATE, ASOF, P_PAID,
           P_BUD, P_FROM, a, P_DATE, ASOF, P_PAID,
           M_AMT, "'Manual Log'!$E$4:$E$500", a, M_CAT, M_DATE, ASOF))
at["K3"] = "=Setup!$B$5"
at["A2"] = '=IF(Setup!$B$9="","","기준일: "&TEXT(Setup!$B$9,"yyyy-mm-dd")&" 까지의 거래 반영")'
at["A2"].font = Font(name="Arial", size=10, italic=True, color=GREEN)

# =====================================================================
# 7.  Calendar — selectable account, Paid?-aware balance, residual column
# =====================================================================
cal = wb["Calendar"]
cal["K3"] = "Main Checking"
cal["K3"].font = Font(name="Arial", size=11, bold=True, color=PURPLE)
cal["L3"] = ('=IF(K3="","",IF(COUNTIF(ActiveAccountNames,K3)=0,'
             '"⚠ Setup Accounts에 없는 계정입니다",""))')
cal["L3"].font = Font(name="Arial", size=10, bold=True, color="FFC00000")
cal["D5"] = "Variable & Other"

def mov(daytest):
    return ('SUMIFS(%s,%s,$K$3,%s,%s,%s,"Yes")+SUMIFS(%s,%s,$K$3,%s,%s,%s,"<>Yes")'
            '-SUMIFS(%s,%s,$K$3,%s,%s,%s,"Yes")-SUMIFS(%s,%s,$K$3,%s,%s,%s,"<>Yes")'
            '+SUMIFS(%s,%s,$K$3,%s,"Income",%s,%s)'
            '-SUMIFS(%s,%s,$K$3,%s,"<>Income",%s,%s)'
            % (P_AMT, P_TO, P_DATE, daytest, P_PAID, P_BUD, P_TO, P_DATE, daytest, P_PAID,
               P_AMT, P_FROM, P_DATE, daytest, P_PAID, P_BUD, P_FROM, P_DATE, daytest, P_PAID,
               M_AMT, "'Manual Log'!$E$4:$E$500", M_CAT, M_DATE, daytest,
               M_AMT, "'Manual Log'!$E$4:$E$500", M_CAT, M_DATE, daytest))

for _r in range(6, 37):
    cal.cell(_r, 8).value = ('=IF($A%d="","",IF(COUNTIF(%s,"<="&$A%d)=0,"PRE-PAY",'
                             '"P"&COUNTIF(%s,"<="&$A%d)))' % (_r, PAYD, _r, PAYD, _r))
cal["I5"] = "Day Net (helper)"
for r in range(6, 37):
    a = "$A%d" % r
    day = [(M_DATE, a)]
    cal.cell(r, 9).value = '=IF(%s="","",%s)' % (a, mov(a))
    cal.cell(r, 9).font = Font(name="Arial", size=10, color="FFBFBFBF")
    cal.cell(r, 2).value = (
        '=IF(%s="","",SUMIFS(%s,%s,%s,%s,"Income",%s,"Yes")+SUMIFS(%s,%s,%s,%s,"Income",%s,"<>Yes")+%s)'
        % (a, P_AMT, P_DATE, a, P_TYPE, P_PAID, P_BUD, P_DATE, a, P_TYPE, P_PAID,
           mlsum(day, (M_CAT, '"Income"'))))
    cal.cell(r, 3).value = (
        '=IF(%s="","",SUMIFS(%s,%s,%s,%s,"<>Income",%s,"<>Transfer",%s,"Yes")'
        '+SUMIFS(%s,%s,%s,%s,"<>Income",%s,"<>Transfer",%s,"<>Yes")'
        '+%s+%s+%s+%s)'
        % (a, P_AMT, P_DATE, a, P_TYPE, P_TYPE, P_PAID,
           P_BUD, P_DATE, a, P_TYPE, P_TYPE, P_PAID,
           mlsum(day, (M_CAT, '"Bills"')), mlsum(day, (M_CAT, '"Subscription"')),
           mlsum(day, (M_CAT, '"Savings"')), mlsum(day, (M_CAT, '"Debt"'))))
    cal.cell(r, 4).value = '=IF(%s="","",%s)' % (a, other_expense(day))
    if r == 6:
        cal.cell(r, 6).value = (
            '=IF(%s="","",IFERROR(LOOKUP(2,1/(%s=$K$3),Setup!$C$%d:$C$%d),0)+%s+$I%d)'
            % (a, ACC_A, ACC_FIRST, ACC_LAST, mov('"<"&%s' % a), r))
    else:
        cal.cell(r, 6).value = '=IF(%s="","",$F%d+$I%d)' % (a, r - 1, r)
cal.column_dimensions["I"].hidden = True

# =====================================================================
# 8.  Monthly Dashboard
# =====================================================================
md = wb["Monthly Dashboard"]
A = "$B$3"
mp, pp = month_period(A), month_period_p(A)
md["B6"] = '=%s+%s' % (paysum(pp, (P_TYPE, '"Income"')), mlsum(mp, (M_CAT, '"Income"')))
md["B7"] = ('=%s+%s+%s'
            % (paysum(pp, (P_TYPE, '"<>Income"'), (P_TYPE, '"<>Savings"'),
                      (P_TYPE, '"<>Debt"'), (P_TYPE, '"<>Transfer"')),
               mlsum(mp, (M_CAT, '"Bills"')), mlsum(mp, (M_CAT, '"Subscription"'))))
md["A8"] = "Variable & Other"
md["B8"] = '=%s' % other_expense(mp)
md["B9"] = '=%s+%s' % (paysum(pp, (P_TYPE, '"Savings"')), mlsum(mp, (M_CAT, '"Savings"')))
md["B10"] = '=%s+%s' % (paysum(pp, (P_TYPE, '"Debt"')), mlsum(mp, (M_CAT, '"Debt"')))
for i in range(6):
    r = 7 + i
    md.cell(r, 6).value = '=IF(_Lists!$AI$%d="","",_Lists!$AI$%d)' % (2 + i, 2 + i)
    md.cell(r, 6).font = Font(name="Arial", size=11, color=GREEN)
    md.cell(r, 7).value = ('=IF($F%d="",0,SUMIFS(%s,%s,$F%d,%s,">="&%s,%s,"<"&EDATE(%s,1)))'
                           % (r, P_BUD, P_SUB, r, P_DATE, A, P_DATE, A))
    md.cell(r, 8).value = ('=IF($F%d="",0,SUMIFS(%s,%s,$F%d,%s,">="&%s,%s,"<"&EDATE(%s,1))'
                           '+SUMIFS(%s,%s,$F%d,%s,">="&%s,%s,"<"&EDATE(%s,1)))'
                           % (r, P_AMT, P_SUB, r, P_DATE, A, P_DATE, A,
                              M_AMT, M_SUB, r, M_DATE, A, M_DATE, A))
    md.cell(r, 9).value = '=G%d-H%d' % (r, r)

# =====================================================================
# 9.  Paycheck Dashboard 1
# =====================================================================
p1 = wb["Paycheck Dashboard 1"]
per = [(P_DATE, '">="&$B$3'), (P_DATE, '"<"&$E$3')]
mper = [(M_DATE, '">="&$B$3'), (M_DATE, '"<"&$E$3')]
p1["C3"] = ('=IF($B$3="","",IF(COUNTIF(' + PAYD + ',$B$3)=0,'
            '"⚠ 급여일 목록에 없는 날짜입니다",'
            'IF(YEAR($B$3)<>Setup!$B$4,"⚠ Budget Year 밖의 날짜입니다","")))')
p1["B6"] = '=%s+%s' % (paysum(per, (P_TYPE, '"Income"')), mlsum(mper, (M_CAT, '"Income"')))
p1["B7"] = ('=%s+%s+%s'
            % (paysum(per, (P_TYPE, '"<>Income"'), (P_TYPE, '"<>Savings"'),
                      (P_TYPE, '"<>Debt"'), (P_TYPE, '"<>Transfer"')),
               mlsum(mper, (M_CAT, '"Bills"')), mlsum(mper, (M_CAT, '"Subscription"'))))
p1["A8"] = "Variable & Other"
p1["B8"] = '=%s' % other_expense(mper)
p1["B9"] = '=%s+%s' % (paysum(per, (P_TYPE, '"Savings"')), mlsum(mper, (M_CAT, '"Savings"')))
p1["B10"] = '=%s+%s' % (paysum(per, (P_TYPE, '"Debt"')), mlsum(mper, (M_CAT, '"Debt"')))
p1["B11"] = "=B6-B7-B8-B9-B10"
p1["B12"] = "=B6-B7-B9-B10"                       # P0-06
p1["A13"] = "남은 일수"
p1["B13"] = '=IF(OR($B$3="",$E$3=""),"",MAX(0,$E$3-MAX($B$3,Setup!$B$9)))'
p1["A13"].font = Font(name="Arial", size=11, bold=False, color=BLACK)
for i in range(31):
    r = 6 + i
    p1.cell(r, 6).value = '=IF(OR($B$3="",$B$3+%d>=$E$3),"",$B$3+%d)' % (i, i)
    p1.cell(r, 7).value = '=IF($F%d="","",%s)' % (r, other_expense([(M_DATE, '$F%d' % r)]))
    p1.cell(r, 8).value = '=IF($F%d="","",SUM($G$6:G%d))' % (r, r)
    p1.cell(r, 9).value = '=IF($F%d="","",MAX(0,$B$12-$H%d)/MAX(1,$E$3-$F%d))' % (r, r, r)

# =====================================================================
# 10. Paycheck Dashboard 2
# =====================================================================
p2 = wb["Paycheck Dashboard 2"]
for r in range(7, 12):
    b = "B%d" % r
    nxt = ('IFERROR(INDEX(%s,MATCH(%s,%s,0)+1),EOMONTH($B$3,0)+1)' % (PAYD, b, PAYD))
    per = [(P_DATE, '">="&%s' % b), (P_DATE, '"<"&%s' % nxt)]
    mper = [(M_DATE, '">="&%s' % b), (M_DATE, '"<"&%s' % nxt)]
    p2.cell(r, 3).value = ('=IF(%s="","",%s+%s)'
                           % (b, paysum(per, (P_TYPE, '"Income"')), mlsum(mper, (M_CAT, '"Income"'))))
    p2.cell(r, 4).value = ('=IF(%s="","",%s+%s+%s)'
                           % (b, paysum(per, (P_TYPE, '"<>Income"'), (P_TYPE, '"<>Savings"'),
                                        (P_TYPE, '"<>Debt"'), (P_TYPE, '"<>Transfer"')),
                              mlsum(mper, (M_CAT, '"Bills"')), mlsum(mper, (M_CAT, '"Subscription"'))))
    p2.cell(r, 5).value = '=IF(%s="","",%s)' % (b, other_expense(mper))
    p2.cell(r, 6).value = ('=IF(%s="","",%s+%s)'
                           % (b, paysum(per, (P_TYPE, '"Savings"')), mlsum(mper, (M_CAT, '"Savings"'))))
    p2.cell(r, 7).value = ('=IF(%s="","",%s+%s)'
                           % (b, paysum(per, (P_TYPE, '"Debt"')), mlsum(mper, (M_CAT, '"Debt"'))))
p2["E6"] = "Variable & Other"

# =====================================================================
# 11. Annual Tracker  (+ yearly total row)
# =====================================================================
an = wb["Annual Tracker"]
an["D4"] = "Variable & Other"
for i in range(12):
    r = 5 + i
    a = "$A%d" % r
    mp, pp = month_period(a), month_period_p(a)
    an.cell(r, 2).value = '=%s+%s' % (paysum(pp, (P_TYPE, '"Income"')), mlsum(mp, (M_CAT, '"Income"')))
    an.cell(r, 3).value = ('=%s+%s+%s'
                           % (paysum(pp, (P_TYPE, '"<>Income"'), (P_TYPE, '"<>Savings"'),
                                     (P_TYPE, '"<>Debt"'), (P_TYPE, '"<>Transfer"')),
                              mlsum(mp, (M_CAT, '"Bills"')), mlsum(mp, (M_CAT, '"Subscription"'))))
    an.cell(r, 4).value = '=%s' % other_expense(mp)
    an.cell(r, 5).value = '=%s+%s' % (paysum(pp, (P_TYPE, '"Savings"')), mlsum(mp, (M_CAT, '"Savings"')))
    an.cell(r, 6).value = '=%s+%s' % (paysum(pp, (P_TYPE, '"Debt"')), mlsum(mp, (M_CAT, '"Debt"')))
an["A17"] = "YEAR TOTAL"
an["A17"].font = Font(name="Arial", size=11, bold=True, color=BLACK)
for c in range(2, 8):
    cell = an.cell(17, c)
    cell.value = "=SUM(%s5:%s16)" % (get_column_letter(c), get_column_letter(c))
    cell._style = copy.copy(an.cell(16, c)._style)
    cell.font = Font(name="Arial", size=11, bold=True, color=BLACK)
    cell.border = Border(top=Side(style="thin", color="FF000000"))
an["A18"] = "MONTHLY AVG"
for c in range(2, 8):
    cell = an.cell(18, c)
    cell.value = "=%s17/12" % get_column_letter(c)
    cell._style = copy.copy(an.cell(16, c)._style)

# =====================================================================
# 12. Distribution / Partner dashboards
# =====================================================================
for sheet, owner in [("Distribution Dashboard", "Joint"),
                     ("Partner 1 Dashboard", "Partner 1"),
                     ("Partner 2 Dashboard", "Partner 2")]:
    d = wb[sheet]
    d["B3"] = _dt.datetime(2026, 4, 1)
    A = "$B$3"
    mp = month_period(A) + [(M_OWN, '"%s"' % owner)]
    pp = month_period_p(A) + [(P_OWN, '"%s"' % owner)]
    d["B6"] = ('=IF(Setup!$B$7="No","",%s+%s)'
               % (paysum(pp, (P_TYPE, '"Income"')), mlsum(mp, (M_CAT, '"Income"'))))
    d["B7"] = ('=IF(Setup!$B$7="No","",%s+%s+%s)'
               % (paysum(pp, (P_TYPE, '"<>Income"'), (P_TYPE, '"<>Variable"'),
                         (P_TYPE, '"<>Transfer"')),
                  mlsum(mp, (M_CAT, '"Bills"')), mlsum(mp, (M_CAT, '"Subscription"'))))
    d["A8"] = "Variable & Other"
    d["B8"] = ('=IF(Setup!$B$7="No","",%s+%s)'
               % (paysum(pp, (P_TYPE, '"Variable"')), other_expense(mp)))

# =====================================================================
# 13. Savings / Debt / Net Worth / 50-30-20
# =====================================================================
sv = wb["Savings-Sinking Funds"]
for r in range(7, 44):
    sv.cell(r, 1).value = None                    # placeholder "Goal 4".."Goal 40"
sv["C4"] = 5000               # portion of the savings account already earmarked
sv["A5"], sv["B5"], sv["C5"], sv["H5"] = "Vacation", 3000, 400, _dt.datetime(2027, 5, 1)
sv["A6"], sv["B6"], sv["C6"], sv["H6"] = "Laptop", 2000, 300, _dt.datetime(2026, 11, 1)
for r in range(4, 44):
    a = "$A%d" % r
    sv.cell(r, 4).value = (
        '=IF(%s="","",SUMIFS(%s,%s,%s,%s,"Savings",%s,"<="&%s)'
        '+SUMIFS(%s,%s,%s,%s,"Savings",%s,"<="&%s))'
        % (a, P_AMT, P_SUB, a, P_TYPE, P_DATE, ASOF,
           M_AMT, M_SUB, a, M_CAT, M_DATE, ASOF))
sv["A45"] = "TOTAL"
sv["A45"].font = Font(name="Arial", size=11, bold=True, color=BLACK)
for c, col in [(2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F")]:
    cell = sv.cell(45, c)
    cell.value = "=SUM(%s4:%s43)" % (col, col)
    cell._style = copy.copy(sv.cell(4, c)._style)
    cell.font = Font(name="Arial", size=11, bold=True, color=BLACK)
    cell.border = Border(top=Side(style="thin", color="FF000000"))
sv["A2"] = '=IF(Setup!$B$9="","","Contributions는 기준일("&TEXT(Setup!$B$9,"yyyy-mm-dd")&")까지 집계. Goal 이름은 Setup Sub-category와 정확히 같아야 합니다")'
sv["A2"].font = Font(name="Arial", size=10, italic=True, color=GREEN)

DEBTS = [("Credit Card", 2400, 0.219, 80, 220),
         ("Car Loan", 12000, 0.059, 350, 0),
         ("Student Loan", 7000, 0.045, 150, 0)]
for name in ("Debt Payoff", "Debt Custom"):
    d = wb[name]
    d["J5"] = "Payoff Order"
    d["J5"]._style = copy.copy(d["I5"]._style)
    for i in range(10):
        r = 6 + i
        if i < len(DEBTS):
            nm, bal, apr, mn, ex = DEBTS[i]
            d.cell(r, 1).value = nm
            d.cell(r, 2).value = bal
            d.cell(r, 3).value = apr
            d.cell(r, 4).value = mn
            d.cell(r, 5).value = ex if name == "Debt Payoff" else 0
        for c in range(1, 11):
            if r > 8:
                d.cell(r, c)._style = copy.copy(d.cell(8, c)._style)
        a = "$A%d" % r
        d.cell(r, 6).value = (
            '=IF(%s="","",MAX(0,$B%d-SUMIFS(%s,%s,%s,%s,"Debt",%s,"<="&%s)'
            '-SUMIFS(%s,%s,%s,%s,"Debt",%s,"<="&%s)))'
            % (a, r, P_AMT, P_SUB, a, P_TYPE, P_DATE, ASOF,
               M_AMT, M_SUB, a, M_CAT, M_DATE, ASOF))
        d.cell(r, 7).value = '=IF(%s="","",$D%d+$E%d)' % (a, r, r)
        d.cell(r, 8).value = (
            '=IF(%s="","",IF(OR($G%d<=0,$F%d<=0),0,'
            'IF($G%d<=$F%d*$C%d/12,"이자 초과",ROUNDUP(NPER($C%d/12,-$G%d,$F%d),0))))'
            % (a, r, r, r, r, r, r, r, r))
        d.cell(r, 9).value = '=IF(OR(%s="",$B%d=0),"",1-$F%d/$B%d)' % (a, r, r, r)
        d.cell(r, 10).value = (
            '=IF(OR(%s="",$F%d<=0),"",'
            'IF($B$3="Snowball",RANK($F%d,$F$6:$F$15,1),'
            'IF($B$3="Credit Score Focus",RANK($F%d,$F$6:$F$15,0),'
            'RANK($C%d,$C$6:$C$15,0))))' % (a, r, r, r, r))
    d["A18"] = ("Current Balance는 Starting Balance 대비 상환 진행도입니다. "
                "카드 신규 사용액은 Accounts Tracker의 카드 잔액에 반영됩니다.")
    d["A18"].font = Font(name="Arial", size=10, italic=True, color=GREEN)
    d["A17"] = ('=IF($B$3="Snowball","Snowball: 잔액이 작은 순서로 상환",'
                'IF($B$3="Credit Score Focus","Credit Score Focus: 잔액이 큰 순서로 상환",'
                'IF($B$3="Custom","Custom: Extra Payment를 직접 배분","Avalanche: APR이 높은 순서로 상환")))')
    d["A17"].font = Font(name="Arial", size=10, italic=True, color=GREEN)
wb["Debt Custom"]["B3"] = "Custom"

nw = wb["Net Worth Tracker"]
for rng in ["A10:F10", "A17:E17"]:
    nw.unmerge_cells(rng)
for r in range(5, 21):
    for c in range(1, 4):
        nw.cell(r, c).value = None
ASSETS = [("Main Checking", "Cash", None), ("Savings", "Cash", None),
          ("Cash", "Cash", None), ("Retirement", "Investments", 15000),
          ("Car", "Property", 12000)]
for i, (nm, cat, val) in enumerate(ASSETS):
    r = 5 + i
    nw.cell(r, 1).value = nm
    nw.cell(r, 2).value = cat
    if val is None:
        nw.cell(r, 3).value = ("=IFERROR(INDEX('Accounts Tracker'!$F$4:$F$22,"
                               "MATCH(A%d,'Accounts Tracker'!$A$4:$A$22,0)),0)" % r)
        nw.cell(r, 3).font = Font(name="Arial", size=11, color=GREEN)
    else:
        nw.cell(r, 3).value = val
        nw.cell(r, 3).font = Font(name="Arial", size=11, color=BLUE)
    for c in range(1, 4):
        nw.cell(r, c)._style = copy.copy(nw.cell(r, c)._style)
for r in range(10, 17):
    nw.cell(r, 3).font = Font(name="Arial", size=11, color=BLUE)
nw["A18"] = "LIABILITIES"
nw["A18"].font = Font(name="Arial", size=12, bold=True, color=BLACK)
nw["A19"], nw["B19"], nw["C19"] = "Liability", "Category", "Balance"
for c in range(1, 4):
    nw.cell(19, c)._style = copy.copy(nw.cell(4, c)._style)
LIABS = [("Credit Card", "Revolving", None), ("Car Loan", "Loan", "debt"),
         ("Student Loan", "Loan", "debt")]
for i, (nm, cat, src) in enumerate(LIABS):
    r = 20 + i
    nw.cell(r, 1).value = nm
    nw.cell(r, 2).value = cat
    if src is None:
        nw.cell(r, 3).value = ("=ABS(IFERROR(INDEX('Accounts Tracker'!$F$4:$F$22,"
                               "MATCH(A%d,'Accounts Tracker'!$A$4:$A$22,0)),0))" % r)
    else:
        nw.cell(r, 3).value = ("=IFERROR(INDEX('Debt Payoff'!$F$6:$F$15,"
                               "MATCH(A%d,'Debt Payoff'!$A$6:$A$15,0)),0)" % r)
    nw.cell(r, 3).font = Font(name="Arial", size=11, color=GREEN)
for r in range(23, 30):
    nw.cell(r, 3).font = Font(name="Arial", size=11, color=BLUE)
nw.merge_cells("A18:F18")
nw.merge_cells("A32:E32")
nw["A32"] = "NET WORTH"
nw["A32"].font = Font(name="Arial", size=12, bold=True, color=BLACK)
nw["A33"], nw["B33"] = "Total Assets", "=SUM(C5:C16)"
nw["A34"], nw["B34"] = "Total Liabilities", "=SUM(C20:C30)"
nw["A35"], nw["B35"] = "Net Worth", "=B33-B34"
for r in (33, 34, 35):
    nw.cell(r, 2)._style = copy.copy(nw.cell(5, 3)._style)
    nw.cell(r, 2).font = Font(name="Arial", size=11, bold=(r == 35), color=BLACK)

fd = wb["50-30-20 Dashboard"]
A = "$B$3"
fd["A10"] = "Take-home Income (기준)"
fd["A10"].font = Font(name="Arial", size=11, bold=True, color=BLACK)
fd["B10"] = ('=%s+%s' % (paysum(month_period_p(A), (P_TYPE, '"Income"')),
                         mlsum(month_period(A), (M_CAT, '"Income"'))))
fd["B10"]._style = copy.copy(fd["C6"]._style)
for i, grp in enumerate(["Needs", "Wants", "Savings"]):
    r = 6 + i
    fd.cell(r, 4).value = '=IF($B$10=0,0,C%d/$B$10)' % r
fd["A12"] = "Actual %는 해당 월 수입 대비 비율입니다."
fd["A12"].font = Font(name="Arial", size=10, italic=True, color=GREEN)

# =====================================================================
# 14. Bonus — dates follow Budget Year, add a summary
# =====================================================================
bn = wb["Bonus"]
for i in range(96):
    r = 5 + i
    bn.cell(r, 1).value = ('=IF(DATE(Setup!$B$4,1,1)+%d>DATE(Setup!$B$4,12,31),"",'
                           'DATE(Setup!$B$4,1,1)+%d)' % (i, i))
    bn.cell(r, 1).number_format = "yyyy-mm-dd"
    bn.cell(r, 1).font = Font(name="Arial", size=11, color=BLACK)
bn["F4"] = "SUMMARY"
bn["F4"].font = Font(name="Arial", size=12, bold=True, color=BLACK)
bn["F5"], bn["G5"] = "No-spend days", '=COUNTIF($B$5:$B$100,"Yes")'
bn["F6"], bn["G6"] = "Logged days", '=COUNTA($B$5:$B$100)'
bn["F7"], bn["G7"] = "No-spend rate", '=IF(G6=0,0,G5/G6)'
bn["G7"].number_format = "0.0%"
bn["F8"] = "표시 기간: Budget Year 1월 1일부터 96일"
bn["F8"].font = Font(name="Arial", size=10, italic=True, color=GREEN)

# =====================================================================
# 15. QA_TEST — keep it honest, add checks for the new wiring
# =====================================================================
qa = wb["QA_TEST"]
qa["C10"] = ('=SUMPRODUCT(--(%s<>""),--(%s<>""),'
             '--(COUNTIFS(%s,%s,%s,%s,%s,"Yes")=0))'
             % (M_CAT, M_SUB, CAT_A, M_CAT, CAT_B, M_SUB, CAT_E))
NEW_TESTS = [
    ("Safe to Spend 분리", "=1",
     "=IF('Paycheck Dashboard 1'!B11='Paycheck Dashboard 1'!B12,0,1)",
     "Projected End Balance와 다른 값이어야 함"),
    ("Annual 합계 일치", "=0",
     "=ROUND('Annual Tracker'!$B$17-(%s+%s),2)"
     % (paysum([(P_DATE, '">="&DATE(Setup!$B$4,1,1)'),
                (P_DATE, '"<="&DATE(Setup!$B$4,12,31)')], (P_TYPE, '"Income"')),
        mlsum([(M_DATE, '">="&DATE(Setup!$B$4,1,1)'),
               (M_DATE, '"<="&DATE(Setup!$B$4,12,31)')], (M_CAT, '"Income"'))),
     "12개월 합 = 연간 원장 합"),
    ("반복거래 상한 준수", "=0",
     "=COUNTA(Setup!$A$%d:$A$100)" % (REC_LAST + 1),
     "Setup 반복거래 %d행 초과 입력 없음" % REC_LAST),
    ("첫 급여일이 1월", "=1",
     '=IF(COUNT(%s)=0,0,IF(MONTH(MIN(%s))=1,1,0))' % (PAYD, PAYD),
     "급여일이 기준일 앞쪽으로도 생성되는지"),
    ("종속목록 생성", "=1",
     '=IF(_Lists!$AC$2="",0,1)',
     "GroupedSub 목록이 채워짐"),
    ("Budget vs Actual 동적", "=1",
     '=IF(\'Monthly Dashboard\'!$F$7="",0,1)',
     "카테고리가 Setup에서 자동 연결"),
]
for i, (nm, exp, act, why) in enumerate(NEW_TESTS):
    r = 12 + i
    qa.cell(r, 1).value = nm
    qa.cell(r, 2).value = exp
    qa.cell(r, 3).value = act
    qa.cell(r, 4).value = '=IF(B%d=C%d,"PASS","FAIL")' % (r, r)
    qa.cell(r, 5).value = why
    for c in range(1, 6):
        qa.cell(r, c)._style = copy.copy(qa.cell(11, c)._style)

# =====================================================================
# 16. number formats — drop the hard-coded "$" (P0-05)
# =====================================================================
MONEY = "#,##0;[RED](#,##0);\\-"
changed_fmt = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            nf = c.number_format
            if nf and "$" in nf:
                c.number_format = MONEY
                changed_fmt += 1
for ws in wb.worksheets:
    cur = {"Accounts Tracker": "K3", "Manual Log": "L3"}.get(ws.title)
print("number formats normalised:", changed_fmt)

# =====================================================================
# 17. data validation — rebuild from scratch on the fixed ranges
# =====================================================================
def dv(ws, formula, sqref, prompt="목록에서 선택하세요. 목록에 없으면 직접 입력할 수 있습니다."):
    v = DataValidation(type="list", formula1=formula, allow_blank=True,
                       showErrorMessage=False, showInputMessage=True)
    v.promptTitle = "선택"
    v.prompt = prompt
    v.errorTitle = "Invalid selection"
    v.error = "목록에서 값을 선택하세요."
    ws.add_data_validation(v)
    v.add(sqref)

for ws in wb.worksheets:
    ws.data_validations.dataValidation = []

dv(setup, "Currency", "B5")
dv(setup, "PayFrequency", "B6")
dv(setup, "YesNo", "B7")
dv(setup, "AccountType", "B%d:B%d" % (ACC_FIRST, ACC_LAST))
dv(setup, "Owner", "E%d:E%d" % (ACC_FIRST, ACC_LAST))
dv(setup, "YesNo", "F%d:F%d" % (ACC_FIRST, ACC_LAST))
dv(setup, "MainCategoryBase", "A%d:A%d" % (CAT_FIRST, CAT_LAST))
dv(setup, "BudgetGroup", "C%d:C%d" % (CAT_FIRST, CAT_LAST))
dv(setup, "Owner", "D%d:D%d" % (CAT_FIRST, CAT_LAST))
dv(setup, "YesNo", "E%d:E%d" % (CAT_FIRST, CAT_LAST))
dv(setup, "RecurringType", "A%d:A%d" % (REC_FIRST, REC_LAST))
dv(setup, "ActiveSubcategoryNames", "B%d:B%d" % (REC_FIRST, REC_LAST))
dv(setup, "Frequency", "D%d:D%d" % (REC_FIRST, REC_LAST))
dv(setup, "ActiveAccountNames", "G%d:H%d" % (REC_FIRST, REC_LAST))
dv(setup, "Owner", "I%d:I%d" % (REC_FIRST, REC_LAST))

dv(pay, "PaidStatus", "G4:G1487")
dv(ml, "MainCategory", "B4:B500")
dv(ml, 'OFFSET(_Lists!$AC$2,IFERROR(MATCH($B4,_Lists!$AB$2:$AB$%d,0),1)-1,0,'
       'MAX(1,COUNTIF(_Lists!$AB$2:$AB$%d,$B4)),1)' % (LAST_L, LAST_L), "C4:C500")
dv(ml, "ActiveAccountNames", "E4:E500")
dv(cal, "MonthList", "B3")
dv(cal, "ActiveAccountNames", "K3")
for s in ["Monthly Dashboard", "Paycheck Dashboard 2", "Distribution Dashboard",
          "Partner 1 Dashboard", "Partner 2 Dashboard", "50-30-20 Dashboard"]:
    dv(wb[s], "MonthList", "B3")
dv(p1, "PayDates", "B3")
dv(wb["Debt Payoff"], "DebtMethod", "B3")
dv(wb["Debt Custom"], "DebtMethod", "B3")
dv(bn, "NoSpend", "B5:B100")
dv(bn, "Mood", "C5:C100")

# =====================================================================
# 18. conditional formatting range fix
# =====================================================================
mlcf = ml.conditional_formatting
for rng in list(mlcf):
    for rule in rng.rules:
        if rule.formula:
            rule.formula = [f.replace("$A$20:$A$60", "$A$%d:$A$%d" % (CAT_FIRST, CAT_LAST))
                             .replace("$B$20:$B$60", "$B$%d:$B$%d" % (CAT_FIRST, CAT_LAST))
                             .replace("$E$20:$E$60", "$E$%d:$E$%d" % (CAT_FIRST, CAT_LAST))
                            for f in rule.formula]

# =====================================================================
# 19. charts
# =====================================================================
from openpyxl.chart import Reference, Series, LineChart
old_chart = p1._charts[0]
newc = LineChart()
newc.title = "Daily Spending"
newc.style = old_chart.style
newc.height, newc.width = 7.5, 15
from openpyxl.chart.series import SeriesLabel
newc.add_data(Reference(p1, min_col=7, max_col=8, min_row=6, max_row=36),
              titles_from_data=False)
newc.set_categories(Reference(p1, min_col=6, min_row=6, max_row=36))
for ser, label in zip(newc.series, ("Daily Spend", "Running Spend")):
    ser.tx = SeriesLabel(v=label)
    ser.smooth = False
p1._charts.clear()
p1.add_chart(newc, "K5")

# =====================================================================
# 20. colour key, hidden sheets, protection, metadata
# =====================================================================
for r in range(4, 44):
    sv.cell(r, 1).font = Font(name="Arial", size=11, color=BLUE)
for r in list(range(5, 17)) + list(range(20, 31)):
    c = nw.cell(r, 3)
    if not (isinstance(c.value, str) and c.value.startswith("=")):
        c.font = Font(name="Arial", size=11, color=BLUE)
for s in ("Debt Payoff", "Debt Custom"):
    wb[s]["B3"].font = Font(name="Arial", size=11, bold=True, color=PURPLE)

wb["_Lists"].sheet_state = "veryHidden"
wb["QA_TEST"].sheet_state = "hidden"

UNLOCK_ANYWAY = {"Payments": ["F4:G1487", "L4:L1487"], "Setup": ["B9"]}
for ws in wb.worksheets:
    if ws.title in ("_Lists",):
        continue
    for row in ws.iter_rows():
        for c in row:
            isf = isinstance(c.value, str) and c.value.startswith("=")
            c.protection = Protection(locked=isf)
    for rng in UNLOCK_ANYWAY.get(ws.title, []):
        sel = ws[rng]
        if not isinstance(sel, tuple):
            sel = ((sel,),)
        for row in sel:
            row = row if isinstance(row, tuple) else (row,)
            for c in row:
                c.protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

wb.properties.creator = "Budget Benchmark"
wb.properties.lastModifiedBy = "Budget Benchmark"
wb.properties.title = "Budget Benchmark V0.5"
wb.properties.subject = "Personal budget workbook"
wb.properties.description = "Setup-driven budget, payments ledger and dashboards."

ins = wb["Instructions"]
ins["A20"] = "V0.5 CHANGES"
ins["A21"] = ("데모 데이터 정리 · 종속 드롭다운 수정 · Setup 유효성 범위 정정 · As of Date 기준일 도입 · "
              "사용자 정의 카테고리 집계 · 통화기호 제거 · Debt 전략/APR 반영 · 수식 경량화 · 시트 보호")
ins["A24"] = "SHEET PROTECTION"
ins["B24"] = "수식 셀은 잠겨 있습니다. 검토 > 시트 보호 해제(암호 없음)로 풀 수 있습니다."

# final sweep: no volatile OFFSET name inside a computed formula
swapped = 0
for _ws in wb.worksheets:
    for _row in _ws.iter_rows():
        for _c in _row:
            v = _c.value
            if isinstance(v, str) and v.startswith("=") and "PayDates" in v:
                _c.value = v.replace("PayDates", PAYD)
                swapped += 1
print("PayDates references de-volatilised:", swapped)

wb.save(DST)
print("saved:", DST)
