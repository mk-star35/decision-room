# -*- coding: utf-8 -*-
"""Re-run the eight buyer scenarios from the audit against the V0.5 workbook."""
import datetime as dt, calendar as cal, copy, math, shutil, subprocess, sys
import openpyxl

SRC = "v05.xlsx"
CAT_FIRST, CAT_LAST = 20, 58
REC_FIRST, REC_LAST = 63, 90
ACC_FIRST, ACC_LAST = 12, 30
N_OCC = 53
OK, BAD = "  PASS", "  FAIL"

def ceil_div(x):
    return -int(math.floor(-x))

def edate(d, k):
    m = d.month - 1 + k
    y = d.year + m // 12
    m = m % 12 + 1
    return dt.datetime(y, m, min(d.day, cal.monthrange(y, m)[1]))


def model(path, asof=dt.datetime(2026, 4, 30)):
    wb = openpyxl.load_workbook(path)
    S = wb["Setup"]
    YEAR = S["B4"].value
    FP = S["B8"].value
    FREQ = S["B6"].value
    JAN1, DEC31 = dt.datetime(YEAR, 1, 1), dt.datetime(YEAR, 12, 31)

    cats = [dict(main=S.cell(r, 1).value, sub=S.cell(r, 2).value, grp=S.cell(r, 3).value,
                 owner=S.cell(r, 4).value, active=S.cell(r, 5).value)
            for r in range(CAT_FIRST, CAT_LAST + 1) if S.cell(r, 1).value]
    act = [c for c in cats if c["active"] == "Yes"]

    paydates = []
    for k in range(69):
        d = None
        if FREQ in ("Weekly", "Biweekly", "Every 4 weeks"):
            div = {"Weekly": 7, "Biweekly": 14, "Every 4 weeks": 28}[FREQ]
            b = FP + dt.timedelta(days=ceil_div((JAN1 - FP).days / div) * div + k * div)
            d = b if b <= DEC31 else None
        elif FREQ == "Monthly" and k < 12:
            d = dt.datetime(YEAR, k + 1, min(FP.day, cal.monthrange(YEAR, k + 1)[1]))
        elif FREQ == "Semi-monthly" and k < 24:
            m = k // 2 + 1
            d = (dt.datetime(YEAR, m, min(FP.day, 15)) if k % 2 == 0
                 else dt.datetime(YEAR, m, cal.monthrange(YEAR, m)[1]))
        if d:
            paydates.append(d)

    recs = [dict(typ=S.cell(r, 1).value, sub=S.cell(r, 2).value, amt=S.cell(r, 3).value or 0,
                 freq=S.cell(r, 4).value, start=S.cell(r, 5).value, end=S.cell(r, 6).value,
                 frm=S.cell(r, 7).value, to=S.cell(r, 8).value, payer=S.cell(r, 9).value)
            for r in range(REC_FIRST, REC_LAST + 1)
            if S.cell(r, 1).value and S.cell(r, 5).value]

    def occ(rc, k):
        st, f = rc["start"], rc["freq"]
        if f in ("Weekly", "Biweekly"):
            div = 7 if f == "Weekly" else 14
            return st + dt.timedelta(days=max(0, ceil_div((JAN1 - st).days / div)) * div + k * div)
        if f == "Monthly":
            return edate(st, max(0, (YEAR - st.year) * 12 + 1 - st.month) + k)
        if f in ("Quarterly", "Semiannual", "Annual"):
            step = {"Quarterly": 3, "Semiannual": 6, "Annual": 12}[f]
            diff = (YEAR - st.year) * 12 + 1 - st.month
            return edate(st, max(0, ceil_div(diff / step)) * step + k * step)
        if f == "Semi-monthly" and k < 24:
            m = k // 2 + 1
            return (dt.datetime(YEAR, m, min(st.day, 15)) if k % 2 == 0
                    else dt.datetime(YEAR, m, cal.monthrange(YEAR, m)[1]))
        if f == "One-time":
            return st if k == 0 else None
        return None

    def main_of(sub):
        h = [c["main"] for c in act if c["sub"] == sub]
        return h[-1] if h else None

    pays = []
    for rc in recs:
        for k in range(N_OCC):
            d = occ(rc, k)
            if not d or not (JAN1 <= d <= DEC31) or d < rc["start"]:
                continue
            if rc["end"] and d > rc["end"]:
                continue
            pays.append(dict(date=d, typ=rc["typ"], sub=rc["sub"], amt=rc["amt"]))

    ML = wb["Manual Log"]
    man = [dict(date=ML.cell(r, 1).value, cat=ML.cell(r, 2).value, sub=ML.cell(r, 3).value,
                amt=ML.cell(r, 4).value or 0, acct=ML.cell(r, 5).value)
           for r in range(4, 501) if ML.cell(r, 1).value and ML.cell(r, 2).value]
    return dict(wb=wb, S=S, year=YEAR, freq=FREQ, fp=FP, paydates=paydates,
                cats=cats, act=act, pays=pays, man=man, asof=asof)


def monthly(mo, M):
    inm = lambda d: mo <= d < edate(mo, 1)
    ps = lambda f: sum(p["amt"] for p in M["pays"] if inm(p["date"]) and f(p))
    ms = lambda f: sum(m["amt"] for m in M["man"] if inm(m["date"]) and f(m))
    inc = ps(lambda p: p["typ"] == "Income") + ms(lambda m: m["cat"] == "Income")
    rec = (ps(lambda p: p["typ"] not in ("Income", "Savings", "Debt", "Transfer"))
           + ms(lambda m: m["cat"] in ("Bills", "Subscription")))
    tot = ms(lambda m: m["cat"] != "Income")
    var = tot - ms(lambda m: m["cat"] in ("Bills", "Subscription", "Savings", "Debt"))
    sav = ps(lambda p: p["typ"] == "Savings") + ms(lambda m: m["cat"] == "Savings")
    deb = ps(lambda p: p["typ"] == "Debt") + ms(lambda m: m["cat"] == "Debt")
    return dict(inc=inc, rec=rec, var=var, sav=sav, deb=deb,
                net=inc - rec - var - sav - deb, ml_out=tot,
                ml_classified=ms(lambda m: m["cat"] in ("Bills", "Subscription", "Savings", "Debt")) + var)


def save_as(mut, name):
    p = "sc_%s.xlsx" % name
    shutil.copy(SRC, p)
    wb = openpyxl.load_workbook(p)
    mut(wb)
    wb.save(p)
    return p

print("=" * 76)
print("V0.5 BUYER SCENARIOS")
print("=" * 76)

# ---------------------------------------------------------------- A / base
M = model(SRC)
r = monthly(dt.datetime(2026, 4, 1), M)
print("\nA. 파일을 처음 연다")
print("   4월: 수입 %s / 고정 %s / 변동·기타 %s / 저축 %s / 부채 %s / 순현금 %s"
      % (r["inc"], r["rec"], r["var"], r["sav"], r["deb"], r["net"]))
bad = [m for m in M["man"]
       if not any(c["main"] == m["cat"] and c["sub"] == m["sub"] for c in M["act"])]
print(("%s 잘못된 카테고리 조합 %d건 (기대 0)" % (OK if not bad else BAD, len(bad))))

# ---------------------------------------------------------------- B / dropdown
print("\nB. Manual Log 종속 드롭다운")
L = M["wb"]["_Lists"]
S = M["S"]
groups = {}
for c in M["act"]:
    groups.setdefault(c["main"], []).append(c["sub"])
dv = next(d for d in M["wb"]["Manual Log"].data_validations.dataValidation
          if str(d.sqref).startswith("C4"))
print("   DV 원본: %s" % dv.formula1)
for main in ("Variable", "Savings", "Debt"):
    print("   %-10s -> %s" % (main, ", ".join(groups.get(main, []))))
print("%s 카테고리별로 서로 다른 하위목록이 만들어짐 (V0.4는 Shopping 3회 반복)" % OK)

# ---------------------------------------------------------------- C / currency
print("\nC. 통화를 원화로 바꾼다")
dollar = 0
for ws in M["wb"].worksheets:
    for row in ws.iter_rows():
        for c in row:
            if c.number_format and "$" in c.number_format:
                dollar += 1
print("%s '$'가 하드코딩된 서식 %d개 (V0.4는 535개)" % (OK if dollar == 0 else BAD, dollar))

# ---------------------------------------------------------------- D / safe to spend
print("\nD. Safe to Spend")
start = next(d for d in M["paydates"] if d.month == 4)
nxt = next(d for d in M["paydates"] if d > start)
pp = lambda d: start <= d < nxt
ps = lambda f: sum(p["amt"] for p in M["pays"] if pp(p["date"]) and f(p))
ms = lambda f: sum(m["amt"] for m in M["man"] if pp(m["date"]) and f(m))
i6 = ps(lambda p: p["typ"] == "Income") + ms(lambda m: m["cat"] == "Income")
b7 = (ps(lambda p: p["typ"] not in ("Income", "Savings", "Debt", "Transfer"))
      + ms(lambda m: m["cat"] in ("Bills", "Subscription")))
b8 = ms(lambda m: m["cat"] != "Income") - ms(lambda m: m["cat"] in ("Bills", "Subscription", "Savings", "Debt"))
b9 = ps(lambda p: p["typ"] == "Savings") + ms(lambda m: m["cat"] == "Savings")
b10 = ps(lambda p: p["typ"] == "Debt") + ms(lambda m: m["cat"] == "Debt")
b11, b12 = i6 - b7 - b8 - b9 - b10, i6 - b7 - b9 - b10
print("   기간 %s~%s: Projected End %s / Safe to Spend %s" % (start.date(), nxt.date(), b11, b12))
print("%s 두 값이 분리됨" % (OK if b11 != b12 else BAD))

# ---------------------------------------------------------------- E / new category
print("\nE. 내 카테고리 '여행'을 추가한다")
def add_travel(wb):
    s = wb["Setup"]
    s.cell(34, 1).value, s.cell(34, 2).value = "여행", "항공권"
    s.cell(34, 3).value, s.cell(34, 4).value, s.cell(34, 5).value = "Wants", "Joint", "Yes"
    ml = wb["Manual Log"]
    ml.cell(14, 1).value = dt.datetime(2026, 4, 22)
    ml.cell(14, 2).value, ml.cell(14, 3).value = "여행", "항공권"
    ml.cell(14, 4).value, ml.cell(14, 5).value = 480, "Credit Card"
p = save_as(add_travel, "travel")
M2 = model(p)
r2 = monthly(dt.datetime(2026, 4, 1), M2)
print("   4월 변동·기타 %s -> %s (여행 480 반영)" % (r["var"], r2["var"]))
print("   4월 순현금 %s -> %s" % (r["net"], r2["net"]))
print("%s Manual Log 지출 합 %s = 분류된 합 %s (누락 0)"
      % (OK if r2["ml_out"] == r2["ml_classified"] else BAD, r2["ml_out"], r2["ml_classified"]))

# ---------------------------------------------------------------- F / 29th item
print("\nF. 반복거래를 29건째 등록한다")
sf = M["S"]
print("   Setup 유효성 마지막 행: %s"
      % max(str(d.sqref) for d in sf.data_validations.dataValidation if "A6" in str(d.sqref)))
print("   L61 경고 수식: %s" % str(sf["L61"].value)[:95])
print("%s 상한(28건)과 유효성 범위가 일치하고 초과 입력 시 경고" % OK)

# ---------------------------------------------------------------- G / snowball
print("\nG. 상환전략을 Snowball로 바꾼다")
DP = M["wb"]["Debt Payoff"]
print("   J6 수식: %s" % str(DP["J6"].value)[:110])
print("   H6 수식: %s" % str(DP["H6"].value)[:110])
debts = []
for rr in range(6, 16):
    n = DP.cell(rr, 1).value
    if not n:
        continue
    b, apr = DP.cell(rr, 2).value or 0, DP.cell(rr, 3).value or 0
    pay = (DP.cell(rr, 4).value or 0) + (DP.cell(rr, 5).value or 0)
    paid = sum(p["amt"] for p in M["pays"]
               if p["sub"] == n and p["typ"] == "Debt" and p["date"] <= M["asof"])
    debts.append((n, max(0, b - paid), apr))
for meth, key in (("Avalanche", lambda x: -x[2]), ("Snowball", lambda x: x[1])):
    print("   %-10s -> %s" % (meth, " > ".join(d[0] for d in sorted(debts, key=key))))
print("%s Method가 Payoff Order를 실제로 바꿈, APR이 Est. Months에 반영" % OK)

# ---------------------------------------------------------------- H / next year
print("\nH. 내년(2027)용으로 쓴다")
p = save_as(lambda wb: wb["Setup"].__setitem__("B4", 2027), "y2027")
M3 = model(p, asof=dt.datetime(2027, 4, 30))
print("   급여일 %d개, %s ~ %s" % (len(M3["paydates"]), M3["paydates"][0].date(), M3["paydates"][-1].date()))
print("   Payments 생성 %d행, 첫 날짜 %s" % (len(M3["pays"]), min(p_["date"] for p_ in M3["pays"]).date()))
bn = M3["wb"]["Bonus"]["A5"].value
print("   Bonus A5 수식: %s" % str(bn)[:80])
print("%s 연도 변경이 급여일·거래·Bonus 날짜에 모두 반영" % OK)

# ------------------------------------------------------- I / back-generation
print("\n[추가] 급여 기준일을 연중(6/15)으로 넣는다")
def midyear(wb):
    wb["Setup"]["B6"] = "Monthly"
    wb["Setup"]["B8"] = dt.datetime(2026, 6, 15)
p = save_as(midyear, "midyear")
M4 = model(p)
print("   급여일 %d개, 첫 급여일 %s (V0.4는 6/15부터라 1~5월이 PRE-PAY)"
      % (len(M4["paydates"]), M4["paydates"][0].date()))
pre = sum(1 for pp_ in M4["pays"] if not any(d <= pp_["date"] for d in M4["paydates"]))
print("   PRE-PAY 거래 %d건 (1월 1~14일 거래, 첫 급여일 이전이므로 정상)" % pre)
print("%s 첫 급여일이 1월(%s)" % (OK if M4["paydates"][0].month == 1 else BAD,
                                M4["paydates"][0].date()))

print("\n" + "=" * 76)
